import openpyxl
import datetime

from typing import Dict, List

#from core.user_card import FromExcelAdditionalUserInfo


def build_dict(user_dict: Dict, user_list: str, key: str, value_index: int):
    try:
        user_dict[key] = user_list[value_index]
    except:
        user_dict[key] = None
    return user_dict
    
def get_players(team_name: str, lst: List):
    player_list = []
    for user in lst:
        user_dict = {
            'Команда': None,
            'Имя': None,
            'Фамилия': None,
            'Дата рождения': None,
            'Позиция': None,
            'Класс': None}
        user_dict['Команда'] = team_name
        user_dict = build_dict(user_dict, user, 'Имя', 0)
        user_dict = build_dict(user_dict, user, 'Фамилия', 1)
        user_dict = build_dict(user_dict, user, 'Дата рождения', 2)
        user_dict = build_dict(user_dict, user, 'Позиция', 3)
        user_dict = build_dict(user_dict, user, 'Класс', 4)
        if user_dict['Имя'] != None:
            player_list.append(user_dict)
    return player_list

def get_coaches(team_name: str, lst: List):
    coach_list = []
    for coach in lst:
        user_coach_dict = {
            'Команда': None,
            'Имя': None,
            'Фамилия': None,
            'Роль': None
            }
        user_coach_dict['Команда'] = team_name
        user_coach_dict = build_dict(user_coach_dict, coach, 'Имя', 0)
        user_coach_dict = build_dict(user_coach_dict, coach, 'Фамилия', 1)
        user_coach_dict = build_dict(user_coach_dict, coach, 'Роль', 3)
        if team_name is not None:
            if user_coach_dict['Имя'] and user_coach_dict['Фамилия'] and user_coach_dict['Роль']:
                coach_list.append(user_coach_dict)
    print(coach_list)
    return coach_list

def parse_team_composition(sheet):  
    lst = []
    team_name = sheet['F1'].value
    for i in range(5, sheet.max_row):
        user_list = []
        for col in sheet.iter_cols(2, sheet.max_column-1):
            value = col[i].value
            if value != 'РУКОВОДИТЕЛИ':
                user_list.append(value)
            else:
                lst.append(1)
        lst.append(user_list)
    to_list_of_dicts(lst, team_name)

def to_list_of_dicts(lst: List, team_name: str):
    index = lst.index(1)
    coaches = get_coaches(team_name, lst[index+3: len(lst)])
    players = get_players(team_name, lst[0 : index])
    return players, coaches


book = openpyxl.load_workbook('parser/Состав команды.xlsx')

for sheet in book:
    parse_team_composition(sheet)
