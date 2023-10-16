import openpyxl
import datetime

from typing import Dict, List

def parse(sheet):
    lst = []
    for i in range(0, sheet.max_row):
        player_list = []
        for col in sheet.iter_cols(1, sheet.max_column):
            value = col[i].value
            player_list.append(value)
        lst.append(player_list)
    result_list = []
    for i in lst:
        player_dict = {
            'Команда': i[0],
            'Имя': i[1].split()[0],
            'Фамилия': i[1].split()[1],
            'Дата рождения': i[2],
            'Класс': i[3],
            'Пересмотр (начало сезона)': i[4]
        }
        result_list.append(player_dict)
    print(result_list)


book = openpyxl.load_workbook('parser/Копия Сводная таблица по командам с классами ЛТ.xlsx')

for i in book:
    if i.title == 'Реестр классификации':
        parse(i)
        
