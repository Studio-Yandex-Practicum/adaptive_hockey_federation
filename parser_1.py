import openpyxl


def parse_team_composition(sheet):  
    results_list = []
    for i in range(0, sheet.max_row):
        row_list = []
        for col in sheet.iter_cols(1, sheet.max_column):
            value = col[i].value
            if value is not None:
                row_list.append(value)
        results_list.append(row_list)
    return results_list


book = openpyxl.load_workbook('Состав команды.xlsx')

for sheet in book:
    print(parse_team_composition(sheet))

book = openpyxl.load_workbook(
    'Копия Сводная таблица по командам с классами ЛТ.xlsx'
    )

for i in book:
    print(parse_team_composition(sheet))