from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

FONT = Font(
    name="Calibri",
    bold=True,
    italic=False,
    vertAlign=None,
    underline="none",
    strike=False,
)

BORDER_LINE = Side(border_style="thin", color="000000")

TITLE_HEIGHT = 25
TITLE_FONT = FONT
TITLE_FONT.size = 13
TITLE_FILL = PatternFill(patternType="solid", fgColor="b4c7dc")

HEADERS_HEIGHT = 20
HEADERS_FONT = FONT
HEADERS_FONT.size = 12
HEADERS_FILL = PatternFill(patternType="solid", fgColor="2a6099")
HEADERS_BORDER = Border(
    top=BORDER_LINE,
    bottom=BORDER_LINE,
    left=BORDER_LINE,
    right=BORDER_LINE
)

ALIGNMENT_CENTER = Alignment(
    horizontal='general',
    vertical='center',
    text_rotation=0,
    wrap_text=False,
    shrink_to_fit=False,
    indent=0
)

ROWS_FILL = PatternFill(patternType="solid", fgColor="dee6ef")
