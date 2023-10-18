from typing import Any, Dict, List

import openpyxl

from adaptive_hockey_federation.core.user_card import (
    ExcelData_2,
    ExcelDataFirstSheet_2,
)

WORKBOOK_PATH = ('adaptive_hockey_federation/parser/'
                 'Копия Сводная таблица по командам с классами ЛТ.xlsx')


def parse_first_sheet(sheet) -> List:
    lst = []
    for i in range(0, sheet.max_row):
        player_list = []
        for col in sheet.iter_cols(1, sheet.max_column):
            value = col[i].value
            player_list.append(value)
        lst.append(player_list)
    return get_players_first_sheet(lst)


def get_players_first_sheet(lst: List) -> List[ExcelDataFirstSheet_2]:
    result_list = []
    for i in lst:
        player_dict = {
            'Команда': i[0],
            'Имя': i[1].split()[0],
            'Фамилия': i[1].split()[1],
            'Дата рождения': i[2],
            'Класс': i[3],
            'Пересмотр (начало сезона)': i[4]
        }
        result_list.append(player_dict)
    return to_list_of_classes_first_sheet(
        result_list
    )


def to_list_of_classes_first_sheet(
        player_list: List[Dict]
) -> List[ExcelDataFirstSheet_2]:
    players = []
    for i in player_list:
        player = ExcelDataFirstSheet_2(
            team=i.get('Команда'),
            name=i.get('Имя'),
            surname=i.get('Фамилия'),
            date_of_birth=i.get('Дата'),
            classification=i.get('Класс'),
            revision=i.get('Пересмотр (начало сезона)')
        )
        players.append(player)
    return players


def parse_sheets(sheet) -> List:
    team_name = sheet['A1'].value
    lst = []
    for i in range(3, sheet.max_row + 1):
        row_list = []
        for j in range(1, sheet.max_column + 1):
            value = sheet.cell(row=i, column=j).value
            row_list.append(value)
        lst.append(row_list)
    return get_players(lst, team_name)


def get_players(
        lst: List,
        team_name: str
) -> List[ExcelData_2]:
    player_list = []
    for i in lst:
        if i[0] is not None:
            if i[3] and i[4] and i[5] and i[6] and i[7]:
                koeff = (i[3] + i[4] + i[5] + i[6] + i[7]) / 5
            else:
                koeff = None
            player_dict = {
                'Команда': team_name,
                'Имя': i[0].split()[0],
                'Фамилия': i[0].split()[1],
                'Дата рождения': i[1],
                'Класс': i[2],
                'Катание лицом вперед': i[3],
                'Катание спиной вперед': i[4],
                'Бросок': i[5],
                'Дриблинг': i[6],
                'Командный хоккей': i[7],
                'ИТОГО': koeff
            }
            player_list.append(player_dict)
    return to_list_of_classes(player_list)


def to_list_of_classes(
        player_list: List[Dict[Any, Any]]
) -> List[ExcelData_2]:
    players = []
    for i in player_list:
        player = ExcelData_2(
            team=i.get('Команда'),
            name=i.get('Имя'),
            surname=i.get('Фамилия'),
            date_of_birth=i.get('Дата'),
            classification=i.get('Класс'),
            riding_face_forward=i.get('Катание лицом вперед'),
            riding_backwards=i.get('Катание спиной вперед'),
            cast=i.get('Бросок'),
            dribbling=i.get('Дриблинг'),
            team_hockey=i.get('Командный хоккей'),
            result=i.get('ИТОГО')
        )
        players.append(player)
    return players


book = openpyxl.load_workbook(WORKBOOK_PATH)


for i in book:
    if i.title == 'Реестр классификации':
        parse_first_sheet(i)
    else:
        parse_sheets(i)
