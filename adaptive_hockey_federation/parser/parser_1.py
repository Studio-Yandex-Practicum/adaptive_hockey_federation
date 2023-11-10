from typing import Any, Dict, List

import openpyxl

from adaptive_hockey_federation.core.user_card import ExcelData  # type: ignore

WORKBOOK_PATH = ('adaptive_hockey_federation/parser/Реестр классов ХДН.xlsx')


def parse(sheet) -> List:
    lst = []
    for i in range(0, 115):
        player_list = []
        for col in sheet.iter_cols(1, sheet.max_column):
            value = col[i].value
            player_list.append(value)
        lst.append(player_list)
    print(lst)
    return get_players(lst)


def get_players(lst: List) -> List[ExcelData]:
    result_list = []
    for i in lst:
        player_dict = {
            'Команда': i[0],
            'Имя': i[1].split()[0],
            'Фамилия': i[1].split()[1],
            'Дата рождения': i[2],
            'Класс': i[3],
            'Пересмотр (начало сезона)': i[4]
        }
        result_list.append(player_dict)
    return to_list_of_classes(
        result_list
    )


def to_list_of_classes(
        player_list: List[Dict[Any, Any]]
) -> List[ExcelData]:
    players = []
    for i in player_list:
        player = ExcelData(
            team=i.get('Команда'),
            name=i.get('Имя'),
            surname=i.get('Фамилия'),
            date_of_birth=i.get('Дата'),
            classification=i.get('Класс'),
        )
        players.append(player)
    return players


book = openpyxl.load_workbook(WORKBOOK_PATH)


for i in book:
    print(parse(i))
