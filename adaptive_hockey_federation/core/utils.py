import os
from datetime import datetime
from typing import Any, List, Optional

from django.conf import settings
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.db.models import QuerySet
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.constants import AgeLimits, FileConstants, TimeFormat
from core.settings.openpyxl_settings import (
    ALIGNMENT_CENTER,
    HEADERS_BORDER,
    HEADERS_FILL,
    HEADERS_FONT,
    HEADERS_HEIGHT,
    ROWS_FILL,
    TITLE_FILL,
    TITLE_FONT,
    TITLE_HEIGHT,
)


def generate_file_name(filename: str, prefix: str) -> str:
    filename, file_extension = os.path.splitext(filename)
    return (
        f"{prefix}-{datetime.now().strftime(TimeFormat.TIME_FORMAT)}"
        f"{file_extension}"
    )


def is_uploaded_file_valid(file: InMemoryUploadedFile) -> bool:
    if (
        file.content_type
        and file.size
        and file.content_type.split("/")[1] in FileConstants.FILE_RESOLUTION
        and file.size <= FileConstants.MAX_UPLOAD_SIZE
    ):
        return True
    return False


def min_date():
    now = datetime.now()
    month_day = format(now.strftime("%m-%d"))
    return f"{str(now.year - AgeLimits.MAX_AGE_PLAYER)}-{month_day}"


def max_date():
    now = datetime.now()
    month_day = format(now.strftime("%m-%d"))
    return f"{str(now.year - AgeLimits.MIN_AGE_PLAYER)}-{month_day}"


def column_width(workbook: Worksheet) -> None:
    for col in workbook.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = max_length + 2
        workbook.column_dimensions[column].width = adjusted_width


def export_excel(
    queryset: QuerySet,
    filename: str,
    title: str,
    excluded_fields: Optional[List[str]] = None,
    fields_order: Optional[List[str]] = None,
) -> str:
    """
    Выгрузка данных в excel (формат xlsx).

    После создания файла возвращает его имя.
    """
    if excluded_fields is None:
        excluded_fields = []

    wb = Workbook()
    del wb["Sheet"]
    ws: Worksheet = wb.create_sheet("Лист1")
    ws.append(["", title])

    if queryset:
        headers, fields = get_fields_and_headers(
            queryset,
            excluded_fields,
            fields_order,
        )
        ws.append(headers)
        add_data_to_worksheet(ws, queryset, fields)

        apply_styles(ws)

    file_path = save_workbook(wb, filename)
    return file_path


def get_fields_and_headers(queryset, excluded_fields, fields_order):
    model_fields = queryset.model._meta.fields
    fields_dict = {
        field.name: str(field.verbose_name)
        for field in model_fields
        if field.name not in excluded_fields
    }

    if fields_order:
        headers = [
            fields_dict[field]
            for field in fields_order
            if field in fields_dict
        ]
        fields = [field for field in fields_order if field in fields_dict]
    else:
        headers = list(fields_dict.values())
        fields = list(fields_dict.keys())

    return headers, fields


def add_data_to_worksheet(ws, queryset, fields):
    for obj in queryset:
        row: List[Any] = []
        for field in fields:
            value = getattr(obj, field)
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            else:
                if hasattr(value, "__str__"):
                    value = value.__str__()

            row.append(value)

        ws.append(row)


def apply_styles(ws):
    column_width(ws)

    ws.row_dimensions[1].height = TITLE_HEIGHT
    ws.row_dimensions[2].height = HEADERS_HEIGHT

    for cell in ws[1]:
        cell.fill = TITLE_FILL
        cell.font = TITLE_FONT
        cell.alignment = ALIGNMENT_CENTER

    for cell in ws[2]:
        cell.fill = HEADERS_FILL
        cell.font = HEADERS_FONT
        cell.alignment = ALIGNMENT_CENTER
        cell.border = HEADERS_BORDER

    number_rows = int(ws.dimensions.split(":")[1][1:])
    for i in range(3, number_rows + 1, 2):
        for cell in ws[i]:
            cell.fill = ROWS_FILL


def save_workbook(wb, filename):
    media_data_path = os.path.join(settings.MEDIA_ROOT, "unloads_data")
    os.makedirs(media_data_path, exist_ok=True)

    timestamp = datetime.now().strftime("%Y.%m.%d_%H%M%S")
    base_filename, file_extension = os.path.splitext(filename)
    filename_with_timestamp = f"{base_filename}_{timestamp}{file_extension}"
    file_path = os.path.join(media_data_path, filename_with_timestamp)
    wb.save(file_path)

    return filename_with_timestamp
